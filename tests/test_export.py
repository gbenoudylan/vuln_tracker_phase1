"""
Test du module d'export (Excel + PDF).

Utilise un DataFrame déjà scoré (simulant la sortie de la Phase 3), pour
tester la génération des rapports indépendamment des appels API NVD/EPSS.
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
from reporting.export import generate_excel_report, generate_pdf_report


def build_sample_scored_df() -> pd.DataFrame:
    return pd.DataFrame({
        "hostname": ["srv-web-01", "srv-db-02", "srv-auth-03", "srv-legacy-04"],
        "cve_id": ["CVE-2024-3400", "CVE-2023-44487", "CVE-2022-0778", "CVE-2021-44228"],
        "vulnerability_name": [
            "PAN-OS Command Injection", "HTTP/2 Rapid Reset",
            "OpenSSL Infinite Loop", "Log4Shell",
        ],
        "cvss_score_final": [10.0, 7.5, 7.5, 10.0],
        "epss_score": [0.99999, 0.99999, 0.73188, 0.94],
        "environment": ["production-internet", "production", "test", "production"],
        "is_eol_flagged": [False, False, True, True],
        "priority_score": [90.0, 75.0, 68.2, 100.0],
        "priority_tier": [
            "P1 - Critique", "P2 - Élevée", "P2 - Élevée", "P1 - Critique",
        ],
    })


def test_generate_excel_report(tmp_path):
    df = build_sample_scored_df()
    output = tmp_path / "test_report.xlsx"
    result_path = generate_excel_report(df, output)

    assert result_path.exists()
    assert result_path.stat().st_size > 0

    # Vérifie que le fichier est un vrai .xlsx lisible
    from openpyxl import load_workbook
    wb = load_workbook(result_path)
    assert "Synthèse" in wb.sheetnames
    assert "Plan de remédiation" in wb.sheetnames

    ws = wb["Plan de remédiation"]
    assert ws.cell(row=1, column=1).value == "Équipement"
    # 4 lignes de données + 1 ligne d'en-tête
    assert ws.max_row == 5

    print(f"OK - test_generate_excel_report ({result_path.stat().st_size} octets)")


def test_generate_pdf_report(tmp_path):
    df = build_sample_scored_df()
    output = tmp_path / "test_report.pdf"
    result_path = generate_pdf_report(df, output)

    assert result_path.exists()
    assert result_path.stat().st_size > 0
    # Vérifie la signature d'un vrai fichier PDF
    with open(result_path, "rb") as f:
        header = f.read(5)
    assert header == b"%PDF-"

    print(f"OK - test_generate_pdf_report ({result_path.stat().st_size} octets)")


def test_export_handles_missing_optional_columns(tmp_path):
    """Un DataFrame sans 'environment' ni 'is_eol_flagged' ne doit pas faire planter l'export."""
    df = pd.DataFrame({
        "hostname": ["srv-01"],
        "cve_id": ["CVE-2024-3400"],
        "cvss_score_final": [9.0],
        "priority_score": [80.0],
        "priority_tier": ["P2 - Élevée"],
    })
    excel_out = tmp_path / "minimal.xlsx"
    pdf_out = tmp_path / "minimal.pdf"

    generate_excel_report(df, excel_out)
    generate_pdf_report(df, pdf_out)

    assert excel_out.exists() and pdf_out.exists()
    print("OK - test_export_handles_missing_optional_columns")


if __name__ == "__main__":
    import tempfile

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)
        test_generate_excel_report(tmp_path)
        test_generate_pdf_report(tmp_path)
        test_export_handles_missing_optional_columns(tmp_path)

    print("\nTous les tests sont passés.")
