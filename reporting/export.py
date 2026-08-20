"""
Module d'export du plan de remédiation.

Prend en entrée le DataFrame final (sortie de la Phase 3 : ingestion +
enrichissement + scoring) et génère deux livrables prêts à diffuser :

    - Un rapport Excel détaillé, avec mise en forme conditionnelle par
      niveau de priorité (pratique pour le suivi opérationnel au quotidien).
    - Un rapport PDF de synthèse, plus adapté à une diffusion managériale
      ou à une preuve d'audit.
"""

from __future__ import annotations
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
)

logger = logging.getLogger("export")

# Couleurs alignées sur les tiers de priorité (cohérent avec le dashboard)
TIER_COLORS_HEX = {
    "P1 - Critique": "D62728",
    "P2 - Élevée": "FF7F0E",
    "P3 - Moyenne": "FFD700",
    "P4 - Faible": "2CA02C",
}

DISPLAY_COLUMNS = [
    ("hostname", "Équipement"),
    ("cve_id", "CVE"),
    ("vulnerability_name", "Vulnérabilité"),
    ("cvss_score_final", "CVSS"),
    ("epss_score", "EPSS"),
    ("environment", "Environnement"),
    ("is_eol_flagged", "EOL"),
    ("priority_score", "Score"),
    ("priority_tier", "Priorité"),
]


def _prepare_export_df(df: pd.DataFrame) -> pd.DataFrame:
    """Sélectionne et renomme les colonnes disponibles pour l'export, triées par priorité."""
    available = [(col, label) for col, label in DISPLAY_COLUMNS if col in df.columns]
    export_df = df[[col for col, _ in available]].copy()
    export_df.columns = [label for _, label in available]
    if "Score" in export_df.columns:
        export_df = export_df.sort_values("Score", ascending=False)
    return export_df


# ---------------------------------------------------------------------
# EXCEL
# ---------------------------------------------------------------------

def generate_excel_report(df: pd.DataFrame, output_path: str | Path) -> Path:
    """
    Génère un rapport Excel avec :
        - un onglet 'Synthèse' (compteurs par tier)
        - un onglet 'Plan de remédiation' (détail trié par priorité,
          coloré par niveau de criticité)
    """
    output_path = Path(output_path)
    export_df = _prepare_export_df(df)

    wb = Workbook()

    # --- Onglet Synthèse ---
    ws_summary = wb.active
    ws_summary.title = "Synthèse"

    ws_summary["A1"] = "Plan de remédiation des vulnérabilités"
    ws_summary["A1"].font = Font(size=14, bold=True)
    ws_summary["A2"] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws_summary["A2"].font = Font(italic=True, color="666666")

    ws_summary["A4"] = "Niveau de priorité"
    ws_summary["B4"] = "Nombre de vulnérabilités"
    ws_summary["A4"].font = ws_summary["B4"].font = Font(bold=True)

    tier_counts = df["priority_tier"].value_counts() if "priority_tier" in df.columns else {}
    row = 5
    for tier, hex_color in TIER_COLORS_HEX.items():
        count = int(tier_counts.get(tier, 0))
        ws_summary[f"A{row}"] = tier
        ws_summary[f"B{row}"] = count
        ws_summary[f"A{row}"].fill = PatternFill("solid", fgColor=hex_color)
        ws_summary[f"A{row}"].font = Font(color="FFFFFF", bold=True)
        row += 1

    ws_summary["A" + str(row + 1)] = "Total"
    ws_summary["B" + str(row + 1)] = len(df)
    ws_summary["A" + str(row + 1)].font = Font(bold=True)

    if "is_eol_flagged" in df.columns:
        n_eol = int(df["is_eol_flagged"].sum())
        ws_summary[f"A{row + 3}"] = "Équipements EOL concernés"
        ws_summary[f"B{row + 3}"] = n_eol

    for col, width in zip("AB", [30, 25]):
        ws_summary.column_dimensions[col].width = width

    # --- Onglet Plan de remédiation ---
    ws = wb.create_sheet("Plan de remédiation")

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(*[Side(style="thin", color="D9D9D9")] * 4)

    for col_idx, col_name in enumerate(export_df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    priority_col_idx = (
        list(export_df.columns).index("Priorité") + 1
        if "Priorité" in export_df.columns else None
    )

    for row_idx, record in enumerate(export_df.to_dict("records"), 2):
        for col_idx, (col_name, value) in enumerate(record.items(), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_name == "EOL" and value:
                cell.font = Font(color="D62728", bold=True)

        if priority_col_idx:
            tier_value = record.get("Priorité")
            hex_color = TIER_COLORS_HEX.get(tier_value)
            if hex_color:
                cell = ws.cell(row=row_idx, column=priority_col_idx)
                cell.fill = PatternFill("solid", fgColor=hex_color)
                cell.font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(export_df.columns, 1):
        max_len = max(
            [len(str(col_name))] + [len(str(v)) for v in export_df[col_name].astype(str)]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(f"Rapport Excel généré : {output_path}")
    return output_path


# ---------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------

def generate_pdf_report(df: pd.DataFrame, output_path: str | Path) -> Path:
    """
    Génère un rapport PDF de synthèse : page de garde avec compteurs,
    puis table des vulnérabilités triées par priorité (P1/P2 en avant).
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(
        str(output_path), pagesize=landscape(A4),
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleCustom", parent=styles["Title"], fontSize=18, spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "Subtitle", parent=styles["Normal"], fontSize=10, textColor=colors.grey,
    )

    elements = [
        Paragraph("Plan de remédiation des vulnérabilités", title_style),
        Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", subtitle_style),
        Spacer(1, 0.6 * cm),
    ]

    # --- Résumé ---
    tier_counts = df["priority_tier"].value_counts() if "priority_tier" in df.columns else {}
    n_eol = int(df["is_eol_flagged"].sum()) if "is_eol_flagged" in df.columns else 0

    summary_data = [["Indicateur", "Valeur"]]
    summary_data.append(["Total vulnérabilités", str(len(df))])
    for tier in TIER_COLORS_HEX:
        summary_data.append([tier, str(int(tier_counts.get(tier, 0)))])
    summary_data.append(["Équipements EOL concernés", str(n_eol)])

    summary_table = Table(summary_data, colWidths=[8 * cm, 4 * cm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 1 * cm))

    # --- Détail (trié par priorité) ---
    export_df = _prepare_export_df(df)
    elements.append(Paragraph("Détail des vulnérabilités (triées par priorité)", styles["Heading2"]))
    elements.append(Spacer(1, 0.3 * cm))

    table_data = [list(export_df.columns)] + export_df.astype(str).values.tolist()
    detail_table = Table(table_data, repeatRows=1)

    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]

    if "Priorité" in export_df.columns:
        priority_col_idx = list(export_df.columns).index("Priorité")
        for row_idx, tier_value in enumerate(export_df["Priorité"], 1):
            hex_color = TIER_COLORS_HEX.get(tier_value)
            if hex_color:
                style_commands.append((
                    "BACKGROUND", (priority_col_idx, row_idx), (priority_col_idx, row_idx),
                    colors.HexColor(f"#{hex_color}"),
                ))
                style_commands.append((
                    "TEXTCOLOR", (priority_col_idx, row_idx), (priority_col_idx, row_idx),
                    colors.white,
                ))

    detail_table.setStyle(TableStyle(style_commands))
    elements.append(detail_table)

    doc.build(elements)
    logger.info(f"Rapport PDF généré : {output_path}")
    return output_path


if __name__ == "__main__":
    import sys
    from pathlib import Path as _Path
    sys.path.insert(0, str(_Path(__file__).parent.parent))

    logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")

    from ingestion.ingest import load_file
    from enrichment.cvss_enrichment import enrich_dataframe
    from scoring.composite_score import run_full_scoring_pipeline

    if len(sys.argv) < 2:
        print("Usage : python -m reporting.export <chemin_fichier>")
        sys.exit(1)

    df_result = run_full_scoring_pipeline(enrich_dataframe(load_file(sys.argv[1])))

    out_dir = _Path(__file__).parent.parent / "output"
    generate_excel_report(df_result, out_dir / "plan_remediation.xlsx")
    generate_pdf_report(df_result, out_dir / "plan_remediation.pdf")
